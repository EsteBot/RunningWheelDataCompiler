# pysimpleGUI for conversion of multiple Running Data Excel files into one Excel file Sheet

# Import required libraries
import PySimpleGUI as sg
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re
from tempfile import TemporaryFile
from pathlib import Path
from datetime import datetime

# validate that the file paths are entered correctly
def is_valid_path(filepath):
    if filepath and Path(filepath).exists():
        return True
    sg.popup_error("A selected file path is incorrect or the field has been left empty.")
    return False

# window appears when the program successfully completes
def nom_window():
    layout = [[sg.Text("\n"
    " All Systems Nominal  \n"
    "\n"
    "")]]
    window = sg.Window((""), layout, modal=True)
    choice = None
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
    window.close()
    
# Define the location of the directory
# Define source file to be analysed 
# Create new Excel file for data to be transfered to 
def running_data_to_Excel_sheet(input_folder, input_sheets, input_filename):
    sheets_list = [x.strip().strip('"') for x in input_sheets.split(",")]
    sheets = sheets_list
    sheet_num = len(sheets)
    print(sheets)
    user_file = input_filename
    # Change the directory
    os.chdir(input_folder)

    # Convert sheet names to rat number IDs
    rat_strings = input_sheets
    rat_integers = []

    string_list = rat_strings
    string_list = string_list.replace("'", "")
    # Split the string based on commas and remove whitespace
    actual_list = [item.strip() for item in string_list.split(",")]

    for string in actual_list:
        match = re.search(r'\d+', string)
        if match:
            integer = int(match.group())
            rat_integers.append(integer)
    
    print(rat_integers)

    # Set working variables
    wheel_circ = 1.081
    bout_list = []
    bout_min  = 0
    bout_count = 0
    prev_cell_value = None
    loop_count = 0
    sheet_loop = 2
    colmn_step = 0
    loop_count1 = 0
    sheet_loop1 = 2
    colmn_step1 = 0
    loop_count2 = 0
    sheet_loop2 = 2
    colmn_step2 = 0
    eshift = -(sheet_num+1)
    sheet_loop3 = 2
    prog_bar_update_val = 0

    input_filename = values["-FILE-"]
            
    # creation of a maximum value for the progress bar function
    input_folder = values["-IN-"]
    
    prog_bar_max_val = 1
    os.chdir(input_folder)
    
    for i in os.listdir():
        prog_bar_max_val += 1
    max = prog_bar_max_val

    # Get the list of files in the directory
    files = os.listdir()

    # Create a new workbook
    newbook = openpyxl.Workbook()
    # Remove the default sheet created with the workbook
    newbook.remove(newbook.active)
    # Add sheet with custum title
    active_sheet = newbook.create_sheet(title='Total Active Data')
    active_hourly_sheet = newbook.create_sheet(title='Hourly Active Data')
    inactive_sheet = newbook.create_sheet(title='Total Inactive Data')

    for filename in os.listdir(input_folder):
    # Check if the file is an Excel file
        # Get a list of all files in the folder
        all_files = os.listdir(input_folder)

        # Filter out only the Excel files (.xlsx)
        # Regular expression pattern for 'MM_DD-YY' format
        date_pattern = r'\d{1,2}[-_]\d{1,2}[-_]\d{2}'

        excel_files = [filename for filename in all_files if filename.endswith('.xlsx') and re.search(date_pattern, filename)]

        # Sort the list of Excel files alphabetically
        def extract_date_from_filename(filename):
            date_string = filename.split()[-1].split('.')[0]  # Extracts the date part from the filename
            return datetime.strptime(date_string, '%m-%d-%y')  # Assuming the date format is MM-DD-YY

        excel_files.sort(key=extract_date_from_filename)

        # Process the files in the sorted order
        for filename in excel_files:
            # Get the full file path
            filepath = os.path.join(input_folder, filename)
            print(filepath)
            wb = openpyxl.load_workbook(filepath)

########### Main Logic for Total Active Data ##################################################

            # Write labels to new Combined Data sheet
            active_sheet.cell(row=1, column=1).value = 'Rat ID'
            # Label rows based on the length of the 'sheets' list
            for index, value in enumerate(rat_integers, start=2):
                active_sheet.cell(row=index, column=1).value = value
            # Label columns by 'Days' based on how many Excel files are looped over 
            # Insert column headers numerically
            for col_idx in range(len(excel_files)):
                # Label rows for bout data
                column_label = 'Day'+ str(col_idx + 1)+'Bouts'
                cell = active_sheet.cell(row=1, column=col_idx + 2)
                cell.value = column_label
                # Label rows for total minutes data
                column_label = 'Day'+ str(col_idx + 1)+'Minutes'
                cell = active_sheet.cell(row=1, column=col_idx + (len(files)) + 2)
                cell.value = column_label
                # Label rows for wheel turn data
                column_label = 'Day'+ str(col_idx + 1)+'Wheel Turns'
                cell = active_sheet.cell(row=1, column=col_idx + ((len(files))*2) + 2)
                cell.value = column_label
                # Label rows for distance (m) data
                column_label = 'Day'+ str(col_idx + 1)+'Distance (m)'
                cell = active_sheet.cell(row=1, column=col_idx + ((len(files))*3)  + 2)
                cell.value = column_label
                # Label rows for average distance per bout data
                column_label = 'Day'+ str(col_idx + 1)+'Distance Per Bout'
                cell = active_sheet.cell(row=1, column=col_idx + ((len(files))*4)  + 2)
                cell.value = column_label
                # Label rows for average bout duration (min) data
                column_label = 'Day'+ str(col_idx + 1)+'Average Bout Duration'
                cell = active_sheet.cell(row=1, column=col_idx + ((len(files))*5)  + 2)
                cell.value = column_label
                # Label rows for speed (m/min) data
                column_label = 'Day'+ str(col_idx + 1)+'Speed (m/min)'
                cell = active_sheet.cell(row=1, column=col_idx + ((len(files))*6)  + 2)
                cell.value = column_label

            # Loop through each sheet and copy the first column's first 3 rows to the new sheet
            for sheet_name in sheets:
                current_sheet = wb[sheet_name]
                # Define the column to iterate over
                column_letter = 'A'
                # Define variables
                wheel_circ = 1.081
                bout_list = []
                bout_min  = 0
                bout_count = 0
                prev_cell_value = None

                # Loop through each cell in the column going down
                #for row in range(2, 3):
                #    cell = current_sheet[column_letter + str(row)]
                #    if cell.value >= 3:
                #       bout_count += 1

                for row in range(2, 722):
                    # Get the cell in the specified column and row
                    cell = current_sheet[column_letter + str(row)]
                    if cell.value is not None and cell.value >= 3:
                        bout_min += 1
                        bout_list.append(cell.value)
                    if cell.value is not None and cell.value >= 3 and prev_cell_value is not None and prev_cell_value < 3:
                        bout_count += 1
                    # Store the current cell value as the previous cell value for the next iteration
                    prev_cell_value = cell.value
    
                bout_sum = sum(bout_list)
                distance = round((bout_sum*wheel_circ),2)
                if bout_count == 0:
                    avg_dist = 0
                else:
                    avg_dist = round((distance/bout_count),2)
                if bout_count == 0:
                    avg_bout_dur = 0
                else:    
                    avg_bout_dur = round((bout_min/bout_count),2)
                if bout_min == 0:
                    speed = 0
                else:
                    speed = round((distance/bout_min),2)
                
                # Write data to new Excel sheet
                #print(loop_count)
                active_sheet.cell(row=sheet_loop+loop_count-colmn_step, column=sheet_loop).value = bout_count
                active_sheet.cell(row=sheet_loop+loop_count-colmn_step, column=(len(files)) + sheet_loop).value = bout_min
                active_sheet.cell(row=sheet_loop+loop_count-colmn_step, column=(len(files))*2+ sheet_loop).value = bout_sum
                active_sheet.cell(row=sheet_loop+loop_count-colmn_step, column=(len(files))*3+ sheet_loop).value = distance
                active_sheet.cell(row=sheet_loop+loop_count-colmn_step, column=(len(files))*4+ sheet_loop).value = avg_dist
                active_sheet.cell(row=sheet_loop+loop_count-colmn_step, column=(len(files))*5+ sheet_loop).value = avg_bout_dur
                active_sheet.cell(row=sheet_loop+loop_count-colmn_step, column=(len(files))*6+ sheet_loop).value = speed

                # Print data summary 
                '''
                print('Total Inactive Data')
                print(f'Number of bouts: {bout_count}')
                print(f'Number of minutes: {bout_min}')
                print(f'Number of wheel turns: {bout_sum}')
                print(f'Distance(m): {distance}')
                print(f'Average distance per bout(m): {avg_dist}')
                print(f'Average bout duration(min): {avg_bout_dur}')
                print(f'Speed(m/min): {speed}')
                print(f'Loop count: {loop_count}')
                print("")
                '''
                loop_count += 1

            sheet_loop += 1
            colmn_step += ((len(sheets))+1)

########### Main Logic for Active Data by Hour ##################################################

            # Write labels to new Combined Data sheet
            active_hourly_sheet.cell(row=1, column=1).value = 'Rat ID'
            # Label rows based on the length of the 'sheets' list
            for index, value in enumerate(rat_integers, start=2):
                active_hourly_sheet.cell(row=index, column=1).value = value

            # Label columns by 'Days' based on how many Excel files are looped over 
            # Insert column headers numerically
            for col_idx in range(len(files) * 12):
                day_idx = col_idx // 12 + 1
                hour_idx = col_idx % 12 + 1

                # Label rows for bout data
                column_label = 'Day{}Hour{} Bouts'.format(day_idx, hour_idx)
                cell = active_hourly_sheet.cell(row=1, column=col_idx + 2)
                cell.value = column_label
                # Label rows for total minutes data
                column_label = 'Day{}Hour{} Minutes'.format(day_idx, hour_idx)
                cell = active_hourly_sheet.cell(row=1, column=(col_idx + (len(files))*12) + 2)
                cell.value = column_label
                # Label rows for wheel turn data
                column_label = 'Day{}Hour{} Wheel Turns'.format(day_idx, hour_idx)
                cell = active_hourly_sheet.cell(row=1, column=(col_idx + ((len(files))*12)*2) + 2)
                cell.value = column_label
                # Label rows for distance (m) data
                column_label = 'Day{}Hour{} Distance(m)'.format(day_idx, hour_idx)
                cell = active_hourly_sheet.cell(row=1, column=(col_idx + ((len(files))*12)*3) + 2)
                cell.value = column_label
                # Label rows for average distance per bout data
                column_label = 'Day{}Hour{} Distance Per Bout'.format(day_idx, hour_idx)
                cell = active_hourly_sheet.cell(row=1, column=(col_idx + ((len(files))*12)*4) + 2)
                cell.value = column_label
                # Label rows for average bout duration (min) data
                column_label = 'Day{}Hour{} Average Bout Duration'.format(day_idx, hour_idx)
                cell = active_hourly_sheet.cell(row=1, column=(col_idx + ((len(files))*12)*5) + 2)
                cell.value = column_label
                # Label rows for speed (m/min) data
                column_label = 'Day{}Hour{} Speed(m/min)'.format(day_idx, hour_idx)
                cell = active_hourly_sheet.cell(row=1, column=(col_idx + ((len(files))*12)*6) + 2)
                cell.value = column_label
            
            start = 2
            end = 62
            step = 60

            for i in range (12):
                eshift += (sheet_num+1)
                # Loop through each sheet and copy the first column's first 3 rows to the new sheet
                for sheet_name in sheets:
                    current_sheet = wb[sheet_name]
                    # Define the column to iterate over
                    column_letter = 'A'
                    # Define variables
                    wheel_circ = 1.081
                    bout_list = []
                    bout_min  = 0
                    bout_count = 0
                    prev_cell_value = None
                        
                    # Get the cell in the specified column and row
                    #print(start)
                    #print(end)
                    for row in range(start, end):
                        cell = current_sheet[column_letter + str(row)]
                        if cell.value is not None and cell.value >= 3:
                            bout_min += 1
                            bout_list.append(cell.value)

                    for row in range(start, end):
                        cell = current_sheet[column_letter + str(row)]
                        if cell.value is not None and cell.value >= 3 and prev_cell_value is not None and prev_cell_value < 3:
                            bout_count += 1
                        # Store the current cell value as the previous cell value for the next iteration
                        prev_cell_value = cell.value
                        
                    bout_sum = sum(bout_list)
                    distance = round((bout_sum*wheel_circ),2)
                    if bout_count == 0:
                        avg_dist = 0
                    else:
                        avg_dist = round((distance/bout_count),2)
                    if bout_count == 0:
                        avg_bout_dur = 0
                    else:    
                        avg_bout_dur = round((bout_min/bout_count),2)
                    if bout_min == 0:
                        speed = 0
                    else:
                        speed = round((distance/bout_min),2)
                    
                    # Write data to new Excel sheet
                    active_hourly_sheet.cell(row=sheet_loop3+loop_count2, column=sheet_loop2).value = bout_count
                    active_hourly_sheet.cell(row=sheet_loop3+loop_count2, column=col_idx + sheet_loop2+1).value = bout_min
                    active_hourly_sheet.cell(row=sheet_loop3+loop_count2, column=(col_idx + ((len(files))*12)) + sheet_loop2+1).value = bout_sum
                    active_hourly_sheet.cell(row=sheet_loop3+loop_count2, column=(col_idx + ((len(files))*12)*2) + sheet_loop2+1).value = distance
                    active_hourly_sheet.cell(row=sheet_loop3+loop_count2, column=(col_idx + ((len(files))*12)*3) + sheet_loop2+1).value = avg_dist
                    active_hourly_sheet.cell(row=sheet_loop3+loop_count2, column=(col_idx + ((len(files))*12)*4) + sheet_loop2+1).value = avg_bout_dur
                    active_hourly_sheet.cell(row=sheet_loop3+loop_count2, column=(col_idx + ((len(files))*12)*5) + sheet_loop2+1).value = speed
                        
                    loop_count2 += 1
            
                start += step
                end += step
                #print(start)
                #print(end)
                sheet_loop2 += 1
                #sheet_loop3 += 1
                loop_count2 = 0

            colmn_step2 += ((len(sheets))+34)
            eshift = -(sheet_num+1)
            sheet_loop3 = 2

########### Main Logic for Total Inactive Data ####################################################

            inactive_sheet.cell(row=1, column=1).value = 'Rat ID'
            # Label rows based on the length of the 'sheets' list
            for index, value in enumerate(rat_integers, start=2):
                inactive_sheet.cell(row=index, column=1).value = value

            # Label columns by 'Days' based on how many Excel files are looped over 
            # Insert column headers numerically
            for col_idx in range(len(files)):
                # Label rows for bout data
                column_label = 'Day'+ str(col_idx + 1)+'Bouts'
                cell = inactive_sheet.cell(row=1, column=col_idx + 2)
                cell.value = column_label
                # Label rows for total minutes data
                column_label = 'Day'+ str(col_idx + 1)+'Minutes'
                cell = inactive_sheet.cell(row=1, column=col_idx + (len(files)) + 2)
                cell.value = column_label
                # Label rows for wheel turn data
                column_label = 'Day'+ str(col_idx + 1)+'Wheel Turns'
                cell = inactive_sheet.cell(row=1, column=col_idx + ((len(files))*2) + 2)
                cell.value = column_label
                # Label rows for distance (m) data
                column_label = 'Day'+ str(col_idx + 1)+'Distance (m)'
                cell = inactive_sheet.cell(row=1, column=col_idx + ((len(files))*3)  + 2)
                cell.value = column_label
                # Label rows for average distance per bout data
                column_label = 'Day'+ str(col_idx + 1)+'Distance Per Bout'
                cell = inactive_sheet.cell(row=1, column=col_idx + ((len(files))*4)  + 2)
                cell.value = column_label
                # Label rows for average bout duration (min) data
                column_label = 'Day'+ str(col_idx + 1)+'Average Bout Duration'
                cell = inactive_sheet.cell(row=1, column=col_idx + ((len(files))*5)  + 2)
                cell.value = column_label
                # Label rows for speed (m/min) data
                column_label = 'Day'+ str(col_idx + 1)+'Speed (m/min)'
                cell = inactive_sheet.cell(row=1, column=col_idx + ((len(files))*6)  + 2)
                cell.value = column_label

            # Loop through each sheet and copy the first column's first 3 rows to the new sheet

            for sheet_name in sheets:
                current_sheet = wb[sheet_name]
                # Define the column to iterate over
                column_letter = 'A'
                # Define variables
                wheel_circ = 1.081
                bout_list = []
                bout_min  = 0
                bout_count = 0
                prev_cell_value = None

                # Loop through each cell in the column going down
                #for row in range(720, 721):
                #    cell = current_sheet[column_letter + str(row)]
                #    if cell.value >= 3:
                #        bout_count += 1

                for row in range(722, 1441):
                    # Get the cell in the specified column and row
                    cell = current_sheet[column_letter + str(row)]
                    if cell.value is not None and cell.value >= 3:
                        bout_min += 1
                        bout_list.append(cell.value)

                    if cell.value is not None and cell.value >= 3 and prev_cell_value is not None and prev_cell_value < 3:
                        bout_count += 1
                    # Store the current cell value as the previous cell value for the next iteration
                    prev_cell_value = cell.value
                    
                bout_sum = sum(bout_list)
                distance = round((bout_sum*wheel_circ),2)
                if bout_count == 0:
                    avg_dist = 0
                else:
                    avg_dist = round((distance/bout_count),2)
                if bout_count == 0:
                    avg_bout_dur = 0
                else:    
                    avg_bout_dur = round((bout_min/bout_count),2)
                if bout_min == 0:
                    speed = 0
                else:
                    speed = round((distance/bout_min),2)
                
                # Write data to new Excel sheet
                inactive_sheet.cell(row=sheet_loop1+loop_count1-colmn_step1, column=sheet_loop1).value = bout_count
                inactive_sheet.cell(row=sheet_loop1+loop_count1-colmn_step1, column=(len(files)) + sheet_loop1).value = bout_min
                inactive_sheet.cell(row=sheet_loop1+loop_count1-colmn_step1, column=(len(files))*2+ sheet_loop1).value = bout_sum
                inactive_sheet.cell(row=sheet_loop1+loop_count1-colmn_step1, column=(len(files))*3+ sheet_loop1).value = distance
                inactive_sheet.cell(row=sheet_loop1+loop_count1-colmn_step1, column=(len(files))*4+ sheet_loop1).value = avg_dist
                inactive_sheet.cell(row=sheet_loop1+loop_count1-colmn_step1, column=(len(files))*5+ sheet_loop1).value = avg_bout_dur
                inactive_sheet.cell(row=sheet_loop1+loop_count1-colmn_step1, column=(len(files))*6+ sheet_loop1).value = speed

                # Print data summary 
                '''
                print('Total Inactive Data')
                print(f'Number of bouts: {bout_count}')
                print(f'Number of minutes: {bout_min}')
                print(f'Number of wheel turns: {bout_sum}')
                print(f'Distance(m): {distance}')
                print(f'Average distance per bout(m): {avg_dist}')
                print(f'Average bout duration(min): {avg_bout_dur}')
                print(f'Speed(m/min): {speed}')
                print(f'Loop count: {loop_count}')
                print("")
                '''
                loop_count1 += 1

            sheet_loop1 += 1
            colmn_step1 += ((len(sheets))+1)

            prog_bar_update_val += 1
            #print("Files Complied: "+str(prog_bar_update_val))

            # records progress by updating prog bar with each file compiled
            window["-Progress_BAR-"].update(max = max, current_count=int(prog_bar_update_val))


        # Save the changes to the Excel file
        test_file = str(user_file)+".xlsx"
        newbook.save(test_file)

        # last prog bar addition indicating the end of the program run
        window["-Progress_BAR-"].update(current_count=int(max))

        # window telling the user the program functioned correctly
        nom_window()   
        break

# creation of a maximum value for the progress bar function
def bar_max(input_folder):
    prog_bar_max_val = 0
    os.chdir(input_folder)
    for i in os.listdir():
        prog_bar_max_val += 1
    #print(prog_bar_max_val)

# main GUI creation and GUI elements
sg.theme('DarkTeal2')

layout = [
    [sg.Text("\n"
             "Select the folder containing the running \n"
             "data Excel files to be processed\n"
             ),
    sg.Input(key="-IN-"),
    sg.FolderBrowse()],

    [sg.Text("Input the name of the new Excel File  \n"
             "compiled running data with be saved to.\n"
             "This file will be created in the current\n"
             "working directory"
             "\n"
             ),
    sg.Input(key="-FILE-")],

    [sg.Text("List all Excel Sheets to be analized     \n"
             "separated by commas\n"
             "Example: rat13, rat14 "
             "\n"
             ),
    sg.Input(key="-SHTS-")],

    [sg.Exit(), sg.Button("Press to analize and compile running data"), 
    sg.Text("\neBot's progress...\n"),
    sg.ProgressBar(20, orientation='horizontal', size=(15,10), 
                border_width=4, bar_color=("Blue", "Grey"),
                key="-Progress_BAR-")]
    
]

# create the window
window = sg.Window("Welcom to eBot's Running Data Analyzer!", layout)

# create an event loop
while True:
    event, values = window.read()
    # end program if user closes window
    if event == "Exit" or event == sg.WIN_CLOSED:
        break
    if event == "Press to analize and compile running data":
        # check file selections are valid
        if (is_valid_path(values["-IN-"])): #and (is_valid_path(values["-OUT-"])):

            running_data_to_Excel_sheet(
            input_folder = values["-IN-"],
            input_filename = values["-FILE-"],
            input_sheets = values["-SHTS-"])

window.close